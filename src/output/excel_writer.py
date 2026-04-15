"""
Generates the multi-tab Excel workbook with analog data, projections,
assumptions, and Corcept enrollment implications.
"""
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.models import (
    DrugProgram, Milestone, MilestoneType, Market, Confidence, Assumption,
    MARKETS, HTA_BODY_NAMES
)


# ── Color scheme ──
FILL_KNOWN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Green
FILL_HIGH = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")      # Light blue
FILL_MODERATE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Light yellow
FILL_LOW = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")       # Light orange
FILL_GAP = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")       # Light red
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # Blue header
FILL_SECTION = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")   # Grey section
FILL_OPTIMISTIC = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_BASE = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
FILL_CONSERVATIVE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_NOTES = Font(name="Calibri", size=8, italic=True, color="666666")
FONT_TITLE = Font(name="Calibri", size=14, bold=True)

ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CONFIDENCE_FILL = {
    Confidence.KNOWN: FILL_KNOWN,
    Confidence.HIGH: FILL_HIGH,
    Confidence.MODERATE: FILL_MODERATE,
    Confidence.LOW: FILL_LOW,
    Confidence.GAP: FILL_GAP,
}

# Milestone display order for analog tabs
ANALOG_MILESTONE_ORDER = [
    ("Trial", [MilestoneType.TRIAL_PCD, MilestoneType.TRIAL_READOUT]),
    ("Regulatory", [
        MilestoneType.FDA_APPROVAL, MilestoneType.EMA_APPROVAL,
        MilestoneType.MHRA_APPROVAL, MilestoneType.HEALTH_CANADA_APPROVAL,
    ]),
    ("Guidelines", [MilestoneType.NCCN_GUIDELINE, MilestoneType.ESMO_GUIDELINE]),
    ("HTA / Reimbursement", [
        MilestoneType.NICE_DECISION, MilestoneType.GBA_DECISION,
        MilestoneType.HAS_DECISION, MilestoneType.AIFA_DECISION,
        MilestoneType.AEMPS_DECISION, MilestoneType.CADTH_DECISION,
    ]),
    ("Commercial", [MilestoneType.COMMERCIAL_LAUNCH]),
    ("SoC Integration", [MilestoneType.SOC_INTEGRATION]),
]

# Which market each milestone primarily applies to
MILESTONE_PRIMARY_MARKET = {
    MilestoneType.TRIAL_PCD: Market.US,
    MilestoneType.TRIAL_READOUT: Market.US,
    MilestoneType.FDA_APPROVAL: Market.US,
    MilestoneType.FDA_SUBMISSION: Market.US,
    MilestoneType.EMA_APPROVAL: Market.DE,  # Pick one EU market
    MilestoneType.MHRA_APPROVAL: Market.UK,
    MilestoneType.HEALTH_CANADA_APPROVAL: Market.CA,
    MilestoneType.NCCN_GUIDELINE: Market.US,
    MilestoneType.ESMO_GUIDELINE: Market.UK,
    MilestoneType.NICE_DECISION: Market.UK,
    MilestoneType.GBA_DECISION: Market.DE,
    MilestoneType.HAS_DECISION: Market.FR,
    MilestoneType.AIFA_DECISION: Market.IT,
    MilestoneType.AEMPS_DECISION: Market.ES,
    MilestoneType.CADTH_DECISION: Market.CA,
    MilestoneType.COMMERCIAL_LAUNCH: Market.US,
    MilestoneType.SOC_INTEGRATION: Market.US,
}


def _format_date(d: Optional[date]) -> str:
    if d is None:
        return "TBD"
    return d.strftime("%b %Y")


def _format_date_quarter(d: Optional[date]) -> str:
    if d is None:
        return "TBD"
    q = (d.month - 1) // 3 + 1
    return f"Q{q} {d.year}"


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _write_analog_tab(ws, analog: DrugProgram):
    """Write a single analog's known milestone data as a milestone x geography matrix."""
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    _set_cell(ws, 1, 1, f"{analog.drug_name} — {analog.indication} ({analog.trial_name})",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT)
    _set_cell(ws, 2, 1, f"Sponsor: {analog.sponsor}", font=FONT_NOTES)

    # Column headers: Milestone | Geography | Date | Lag (months) | Confidence | Source | Notes
    headers = ["Category", "Milestone", "Geography", "Date", "Lag from FDA (months)",
               "Confidence", "Source", "Notes"]
    for col, header in enumerate(headers, 1):
        _set_cell(ws, 4, col, header, font=FONT_HEADER, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    fda_date = analog.get_fda_approval_date()
    row = 5

    for category, mtypes in ANALOG_MILESTONE_ORDER:
        for mtype in mtypes:
            # Find all milestones of this type across markets
            matching = [m for m in analog.milestones if m.milestone_type == mtype]
            if not matching:
                # Still show the row as a gap
                _set_cell(ws, row, 1, category, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 2, mtype.value, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 3, "All", font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 4, "No data", font=FONT_BODY, fill=FILL_GAP,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
                row += 1
                continue

            for m in matching:
                _set_cell(ws, row, 1, category, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 2, mtype.value, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 3, m.market.value, font=FONT_BODY,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

                date_str = _format_date(m.date_value)
                fill = CONFIDENCE_FILL.get(m.confidence, None)
                _set_cell(ws, row, 4, date_str, font=FONT_BODY, fill=fill,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

                # Lag from FDA
                lag_str = ""
                if fda_date and m.date_value and mtype != MilestoneType.FDA_APPROVAL:
                    from src.projection.lag_calculator import months_between
                    lag = months_between(fda_date, m.date_value)
                    lag_str = f"{lag:.1f}"
                _set_cell(ws, row, 5, lag_str, font=FONT_BODY,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

                _set_cell(ws, row, 6, m.confidence.value, font=FONT_BODY,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
                _set_cell(ws, row, 7, m.source, font=FONT_NOTES,
                          alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                _set_cell(ws, row, 8, m.notes, font=FONT_NOTES,
                          alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                row += 1

    # Column widths
    widths = [14, 28, 12, 12, 18, 12, 35, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_derived_lags_tab(ws, all_lags):
    """Write the derived lag analysis showing both analogs side by side."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _set_cell(ws, 1, 1, "Derived Lag Analysis (months from FDA approval)",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT)

    headers = ["Milestone", "Geography", "Enhertu Lag (months)", "Trodelvy 3L+ Lag (months)",
               "Optimistic", "Base (Mid)", "Conservative", "Notes"]
    for col, header in enumerate(headers, 1):
        _set_cell(ws, 3, col, header, font=FONT_HEADER, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    from src.projection.lag_calculator import get_lag_range, DOWNSTREAM_MILESTONES

    row = 4
    for mtype, markets in DOWNSTREAM_MILESTONES:
        for market in markets:
            _set_cell(ws, row, 1, mtype.value, font=FONT_BODY, border=THIN_BORDER)
            _set_cell(ws, row, 2, market.value, font=FONT_BODY,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

            # Find individual analog lags
            enhertu_lag = None
            trodelvy_lag = None
            for lag in all_lags:
                if lag.milestone_type == mtype and lag.market == market:
                    if "Enhertu" in lag.analog_name:
                        enhertu_lag = lag.lag_months
                    elif "Trodelvy" in lag.analog_name:
                        trodelvy_lag = lag.lag_months

            _set_cell(ws, row, 3,
                      f"{enhertu_lag:.1f}" if enhertu_lag is not None else "N/A",
                      font=FONT_BODY, fill=FILL_OPTIMISTIC,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, 4,
                      f"{trodelvy_lag:.1f}" if trodelvy_lag is not None else "N/A",
                      font=FONT_BODY, fill=FILL_CONSERVATIVE,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

            opt, mid, cons = get_lag_range(all_lags, mtype, market)
            _set_cell(ws, row, 5,
                      f"{opt:.1f}" if opt is not None else "N/A",
                      font=FONT_BODY, fill=FILL_OPTIMISTIC,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, 6,
                      f"{mid:.1f}" if mid is not None else "N/A",
                      font=FONT_BODY, fill=FILL_BASE,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, 7,
                      f"{cons:.1f}" if cons is not None else "N/A",
                      font=FONT_BODY, fill=FILL_CONSERVATIVE,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

            notes = ""
            if enhertu_lag is None and trodelvy_lag is None:
                notes = "GAP: No analog data for this milestone/market"
            elif enhertu_lag is None or trodelvy_lag is None:
                notes = "Single analog only — wider uncertainty"
            _set_cell(ws, row, 8, notes, font=FONT_NOTES,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
            row += 1

    widths = [28, 12, 20, 22, 12, 12, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_projection_tab(ws, label: str, scenarios: dict[str, DrugProgram]):
    """Write projection tab with 3 scenarios side by side for a specific drug/indication."""
    base_program = scenarios.get("base") or list(scenarios.values())[0]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    _set_cell(ws, 1, 1,
              f"{base_program.drug_name} — {base_program.indication} ({base_program.trial_name})",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT)
    _set_cell(ws, 2, 1, f"Sponsor: {base_program.sponsor}", font=FONT_NOTES)

    # Headers
    headers = [
        "Category", "Milestone", "Geography",
        "Optimistic", "Base", "Conservative",
        "Confidence", "Source / Basis", "Notes"
    ]
    for col, header in enumerate(headers, 1):
        fill = FILL_HEADER
        if header == "Optimistic":
            fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        elif header == "Base":
            fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        elif header == "Conservative":
            fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        _set_cell(ws, 4, col, header, font=FONT_HEADER, fill=fill,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    row = 5
    for category, mtypes in ANALOG_MILESTONE_ORDER:
        for mtype in mtypes:
            # Collect all markets that have data for this milestone type
            all_markets_with_data = set()
            for scenario in scenarios.values():
                for m in scenario.milestones:
                    if m.milestone_type == mtype:
                        all_markets_with_data.add(m.market)

            if not all_markets_with_data:
                continue

            for market in MARKETS:
                if market not in all_markets_with_data:
                    continue

                _set_cell(ws, row, 1, category, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 2, mtype.value, font=FONT_BODY, border=THIN_BORDER)
                _set_cell(ws, row, 3, market.value, font=FONT_BODY,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

                # Three scenarios
                for i, scenario_name in enumerate(["optimistic", "base", "conservative"]):
                    prog = scenarios.get(scenario_name)
                    m = prog.get_milestone(mtype, market) if prog else None
                    fills = [FILL_OPTIMISTIC, FILL_BASE, FILL_CONSERVATIVE]
                    date_str = _format_date(m.date_value) if m else "TBD"
                    _set_cell(ws, row, 4 + i, date_str, font=FONT_BODY,
                              fill=fills[i], alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

                # Confidence and notes from base scenario
                base_m = scenarios["base"].get_milestone(mtype, market) if "base" in scenarios else None
                if base_m:
                    conf_fill = CONFIDENCE_FILL.get(base_m.confidence)
                    _set_cell(ws, row, 7, base_m.confidence.value, font=FONT_BODY,
                              fill=conf_fill, alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
                    _set_cell(ws, row, 8, base_m.source, font=FONT_NOTES,
                              alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                    _set_cell(ws, row, 9, base_m.notes, font=FONT_NOTES,
                              alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                else:
                    _set_cell(ws, row, 7, "", border=THIN_BORDER)
                    _set_cell(ws, row, 8, "", border=THIN_BORDER)
                    _set_cell(ws, row, 9, "", border=THIN_BORDER)

                row += 1

    widths = [14, 28, 12, 14, 14, 14, 12, 40, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_assumptions_tab(ws, assumptions: list[Assumption]):
    """Write the assumptions and data gaps log."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _set_cell(ws, 1, 1, "Assumptions & Data Gaps Log", font=FONT_TITLE)

    headers = ["ID", "Category", "Description", "Impact", "Status",
               "Related Drug", "Data Gap?", "Action Required"]
    for col, header in enumerate(headers, 1):
        _set_cell(ws, 3, col, header, font=FONT_HEADER, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    for i, a in enumerate(assumptions, start=4):
        _set_cell(ws, i, 1, a.id, font=FONT_BODY, border=THIN_BORDER)
        _set_cell(ws, i, 2, a.category, font=FONT_BODY, border=THIN_BORDER)
        _set_cell(ws, i, 3, a.description, font=FONT_BODY,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        _set_cell(ws, i, 4, a.impact, font=FONT_BODY,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, i, 5, a.status, font=FONT_BODY,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, i, 6, a.related_drug or "General", font=FONT_BODY,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

        gap_str = "YES" if a.gap_flag else "No"
        gap_fill = FILL_GAP if a.gap_flag else None
        _set_cell(ws, i, 7, gap_str, font=FONT_BODY, fill=gap_fill,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

        action = "Requires primary research" if a.gap_flag else ""
        _set_cell(ws, i, 8, action, font=FONT_NOTES,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)

    widths = [8, 16, 70, 10, 10, 14, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_enrollment_implications_tab(
    ws,
    all_program_scenarios: list[tuple[str, dict[str, DrugProgram]]],
    current_date: date = date(2026, 4, 15),
):
    """
    Corcept trial enrollment implications tab.
    Maps SoC integration dates to enrollment window recommendations.
    """
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _set_cell(ws, 1, 1, "Corcept Phase 3 Enrollment Implications",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT)
    _set_cell(ws, 2, 1,
              f"Analysis date: {current_date.strftime('%d %b %Y')}. "
              "SoC integration = guideline inclusion + national reimbursement/access.",
              font=FONT_NOTES)

    # Headers
    row = 4
    _set_cell(ws, row, 1, "Geography", font=FONT_HEADER, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    _set_cell(ws, row, 2, "HTA Body", font=FONT_HEADER, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    col = 3
    drug_col_map = {}
    for label, scenarios in all_program_scenarios:
        drug_col_map[label] = col
        ws.merge_cells(start_row=row - 1, start_column=col, end_row=row - 1, end_column=col + 1)
        _set_cell(ws, row - 1, col, label, font=FONT_SECTION, fill=FILL_SECTION,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, col, "SoC Date (Base)", font=FONT_HEADER, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, col + 1, "Enrollment Window", font=FONT_HEADER, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        col += 2

    # Data rows
    row = 5
    for market in MARKETS:
        _set_cell(ws, row, 1, market.value, font=FONT_BODY,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, 2, HTA_BODY_NAMES.get(market, ""), font=FONT_BODY,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)

        for label, scenarios in all_program_scenarios:
            col = drug_col_map[label]
            base = scenarios.get("base")
            soc_m = base.get_milestone(MilestoneType.SOC_INTEGRATION, market) if base else None

            if soc_m and soc_m.date_value:
                soc_date = soc_m.date_value
                date_str = _format_date_quarter(soc_date)
                months_to_soc = (soc_date.year - current_date.year) * 12 + \
                                (soc_date.month - current_date.month)

                if months_to_soc <= 0:
                    window = "SoC already integrated"
                    fill = FILL_GAP
                elif months_to_soc <= 12:
                    window = f"Narrow ({months_to_soc}mo) — prioritize now"
                    fill = FILL_LOW
                elif months_to_soc <= 24:
                    window = f"Moderate ({months_to_soc}mo) — plan enrollment"
                    fill = FILL_MODERATE
                else:
                    window = f"Extended ({months_to_soc}mo) — flexible timing"
                    fill = FILL_KNOWN
            else:
                date_str = "TBD"
                window = "Cannot assess — resolve data gaps"
                fill = FILL_GAP

            _set_cell(ws, row, col, date_str, font=FONT_BODY,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, col + 1, window, font=FONT_BODY, fill=fill,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)

        row += 1

    # Summary / recommendation
    row += 2
    _set_cell(ws, row, 1, "Key Findings", font=FONT_SECTION)
    row += 1
    recommendations = [
        "US: NCCN already includes both Trodelvy and Dato-DXd pre-approval. SoC integration will occur at commercial launch (H2 2026).",
        "Germany: Access available at EMA approval (free pricing period). Among earliest EU5 markets for SoC shift.",
        "UK: NICE CDF pathway likely for newer ADCs. Full SoC integration may lag 12-28 months post-FDA.",
        "France/Italy/Spain: Longest reimbursement timelines (12-27 months post-FDA). Extended enrollment windows.",
        "Canada: Moderate timeline. CADTH recommendation typically 12-20 months post-FDA.",
        "Dato-DXd CPS>=10 (TROPION-Breast05): Significantly behind — PCD ~Oct 2027, FDA ~2029. Extended window in all markets.",
        "Implication: Geographies with longer HTA timelines (IT, ES, FR) offer wider pre-SoC enrollment windows for Corcept's Phase 3.",
    ]
    for rec in recommendations:
        _set_cell(ws, row, 1, rec, font=FONT_BODY, alignment=ALIGNMENT_LEFT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col - 1)
        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    for c in range(3, col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


def create_workbook(
    analogs: list[DrugProgram],
    all_lags: list,
    projection_sets: list[tuple[str, dict[str, DrugProgram]]],
    assumptions: list[Assumption],
    output_path: str,
):
    """
    Create the complete multi-tab Excel workbook.

    Args:
        analogs: List of analog DrugPrograms (Enhertu, Trodelvy 3L+)
        all_lags: Combined AnalogLag list from both analogs
        projection_sets: List of (label, {scenario_name: DrugProgram}) for each target drug
        assumptions: Full assumptions registry
        output_path: Where to save the .xlsx file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Tab 1 & 2: Analog data
    for analog in analogs:
        short_name = analog.drug_name.split("(")[0].strip()
        ws = wb.create_sheet(title=f"Analog - {short_name}")
        _write_analog_tab(ws, analog)

    # Tab 3: Derived lags
    ws = wb.create_sheet(title="Derived Lags")
    _write_derived_lags_tab(ws, all_lags)

    # Tab 4-7: Projection tabs
    for label, scenarios in projection_sets:
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = label[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_projection_tab(ws, label, scenarios)

    # Tab 8: Assumptions
    ws = wb.create_sheet(title="Assumptions & Gaps")
    _write_assumptions_tab(ws, assumptions)

    # Tab 9: Enrollment Implications
    ws = wb.create_sheet(title="Enrollment Implications")
    _write_enrollment_implications_tab(ws, projection_sets)

    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")
